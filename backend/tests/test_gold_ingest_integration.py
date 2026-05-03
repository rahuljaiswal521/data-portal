"""End-to-end smoke test for the Gold ingest API.

Builds an xlsx in-memory, POSTs it to /preview, then POSTs the IR to /commit,
finally verifies the mart appears in /marts.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.config import settings
from app.dependencies import (
    get_gold_config_service,
    get_gold_ingest_service,
)
from app.main import app


@pytest.fixture
def client(tmp_path: Path, monkeypatch):
    # Redirect gold marts dir to a tmp location so the test is isolated
    monkeypatch.setattr(settings, "gold_conf_dir", str(tmp_path))
    (tmp_path / "marts").mkdir(parents=True, exist_ok=True)
    (tmp_path / "environments").mkdir(parents=True, exist_ok=True)

    # Reset DI cache so the new path is picked up
    get_gold_config_service.cache_clear()
    get_gold_ingest_service.cache_clear()

    return TestClient(app)


def _build_xlsx() -> bytes:
    wb = Workbook()
    # default sheet -> mart
    ws_mart = wb.active
    ws_mart.title = "mart"
    ws_mart.append(["name", "description", "schema", "owner"])
    ws_mart.append(["sales", "Sales analytics mart", "gld_sales", "data-team"])

    ws_dim = wb.create_sheet("dimensions")
    ws_dim.append(
        [
            "dim_name", "description", "source_entity", "business_key",
            "scd_type", "is_conformed", "attribute",
            "attribute_source_column", "attribute_description",
        ]
    )
    ws_dim.append(["dim_customer", "Customer", "slv_customer.customer", "customer_id", "scd2", "true", "customer_name", "", ""])
    ws_dim.append(["dim_customer", "", "", "", "", "", "country_code", "", ""])
    ws_dim.append(["dim_product", "Product", "slv_sales.product", "product_id", "scd2", "false", "product_name", "", ""])

    ws_fact = wb.create_sheet("facts")
    ws_fact.append(
        [
            "fact_name", "description", "source_entity", "grain",
            "load_type", "watermark", "kind", "name",
            "expression", "sk_column", "dim", "source_column",
        ]
    )
    ws_fact.append(["fact_orders", "Orders", "slv_sales.order_line", "order_id, order_line_id", "merge", "order_updated_at", "fk", "customer_fk", "", "customer_sk", "dim_customer", "customer_id"])
    ws_fact.append(["fact_orders", "", "", "", "", "", "fk", "product_fk", "", "product_sk", "dim_product", "product_id"])
    ws_fact.append(["fact_orders", "", "", "", "", "", "measure", "order_amount", "quantity * unit_price", "", "", ""])

    ws_m = wb.create_sheet("metrics")
    ws_m.append(["metric_name", "fact", "formula", "grain", "materialization"])
    ws_m.append(["total_revenue", "fact_orders", "SUM(order_amount)", "order_date", "view"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_preview_then_commit_then_list(client: TestClient):
    xlsx = _build_xlsx()

    # 1. Preview
    r = client.post(
        "/api/v1/gold/ingest/preview",
        files={"file": ("rules.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"default_mart_name": "sales"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"] == {"n_dimensions": 2, "n_facts": 1, "n_metrics": 1}
    assert body["diff"]["exists"] is False
    ir = body["ir"]
    assert ir["mart"]["name"] == "sales"

    # 2. Commit
    r = client.post(
        "/api/v1/gold/ingest/commit",
        json={"ir": ir, "overwrite": False},
    )
    assert r.status_code == 200, r.text
    assert r.json()["mart_name"] == "sales"

    # 3. List
    r = client.get("/api/v1/gold/marts")
    assert r.status_code == 200, r.text
    marts = r.json()
    assert len(marts) == 1
    assert marts[0]["name"] == "sales"
    assert marts[0]["n_dimensions"] == 2
    assert marts[0]["n_facts"] == 1
    assert marts[0]["n_metrics"] == 1

    # 4. Get detail
    r = client.get("/api/v1/gold/marts/sales")
    assert r.status_code == 200, r.text
    detail = r.json()
    assert {d["name"] for d in detail["dimensions"]} == {"dim_customer", "dim_product"}

    # 5. Re-commit without overwrite -> 400
    r = client.post(
        "/api/v1/gold/ingest/commit",
        json={"ir": ir, "overwrite": False},
    )
    assert r.status_code == 400

    # 6. Re-commit with overwrite -> 200
    r = client.post(
        "/api/v1/gold/ingest/commit",
        json={"ir": ir, "overwrite": True},
    )
    assert r.status_code == 200

    # 7. Delete
    r = client.delete("/api/v1/gold/marts/sales")
    assert r.status_code == 204
    r = client.get("/api/v1/gold/marts")
    assert r.json() == []
